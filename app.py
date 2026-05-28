from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, shutil, io

app = Flask(__name__)

DATA_DIR     = os.path.join(os.path.dirname(__file__), "data")
EXCEL_FILE   = os.path.join(DATA_DIR, "truck_trips.xlsx")
SETTINGS_FILE= os.path.join(DATA_DIR, "settings.json")
BACKUP_DIR   = os.path.join(DATA_DIR, "backups")

HEADERS = [
    "Trip No", "Date", "Truck No", "Driver Name",
    "Loading Point", "Delivery Point", "Weight (Tons)",
    "Freight Amount", "Toll Charges", "Commission Charge",
    "Fuel Liters", "Fuel Amount", "Per Trip Expenses",
    "Advance Amount", "Bill Amount", "Total Trip Amount", "Balance Amount"
]

DEFAULT_SETTINGS = {
    "company_name": "GOODS CARRIER",
    "company_address": "",
    "company_phone": "",
    "trucks": [],
    "drivers": []
}

# ── Init ──────────────────────────────────────
os.makedirs(DATA_DIR, exist_ok=True)

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                s = json.load(f)
            for k, v in DEFAULT_SETTINGS.items():
                s.setdefault(k, v)
            return s
        except Exception:
            pass
    return dict(DEFAULT_SETTINGS)

def save_settings_file(settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f, indent=2)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trips"
        hf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        wb.save(EXCEL_FILE)

init_excel()

def load_trips():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None for c in row):
            r = list(row)
            # pad old rows to 17 cols
            while len(r) < 17:
                r.append(None)
            rows.append(r)
    return rows

def save_trip_excel(data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    trip_no = ws.max_row
    ws.append([trip_no] + data)
    last = ws.max_row
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col in range(1, len(HEADERS)+1):
        cell = ws.cell(row=last, column=col)
        if last % 2 == 0: cell.fill = alt
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_FILE)
    return trip_no

def update_trip_excel(trip_no, data):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == trip_no:
            full = [trip_no] + data
            for col, val in enumerate(full, 1):
                ws.cell(row=row[0].row, column=col, value=val)
            break
    wb.save(EXCEL_FILE)

def delete_trip_excel(trip_no):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == trip_no:
            ws.delete_rows(row[0].row)
            break
    wb.save(EXCEL_FILE)

def build_report_wb(trips, title_suffix=""):
    settings = load_settings()
    company  = settings.get("company_name", "GOODS CARRIER")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_col_letter = get_column_letter(len(HEADERS))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"{company} - TRIP SUMMARY REPORT {title_suffix}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 35

    alt = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    totals = [0]*10
    for i, trip in enumerate(trips, 1):
        rn = i + 4
        for col, val in enumerate(trip, 1):
            cell = ws.cell(row=rn, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0: cell.fill = alt
        try:
            for idx, ti in enumerate([7,8,9,10,11,12,13,14,15,16]):
                totals[idx] += float(trip[ti] or 0) if len(trip) > ti else 0
        except (TypeError, ValueError): pass

    tr = len(trips) + 5
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    tfill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    for col, val in zip(range(8, 18), totals):
        cell = ws.cell(row=tr, column=col, value=round(val, 2))
        cell.font = Font(bold=True); cell.fill = tfill
        cell.border = border; cell.alignment = Alignment(horizontal="center")

    col_widths = [8,12,12,15,18,18,12,15,12,16,11,12,15,14,13,16,15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb

# ── Routes ────────────────────────────────────
@app.route("/")
def index():
    return redirect(url_for("trips_view"))

@app.route("/trips")
def trips_view():
    settings = load_settings()
    trips = load_trips()
    trucks  = sorted(set(str(t[2]) for t in trips if t[2]))
    drivers = sorted(set(str(t[3]) for t in trips if t[3]))

    search    = request.args.get("search","").lower().strip()
    truck_f   = request.args.get("truck","All")
    driver_f  = request.args.get("driver","All")
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to","")

    filtered = []
    for t in trips:
        if search and not any(search in str(v).lower() for v in t if v): continue
        if truck_f != "All" and str(t[2]).upper() != truck_f.upper(): continue
        if driver_f != "All" and str(t[3]).lower() != driver_f.lower(): continue
        if date_from and date_to:
            try:
                td = datetime.strptime(str(t[1]), "%d-%m-%Y").date()
                fd = datetime.strptime(date_from, "%Y-%m-%d").date()
                tod= datetime.strptime(date_to,   "%Y-%m-%d").date()
                if not (fd <= td <= tod): continue
            except: pass
        filtered.append(t)

    total_freight = sum(float(t[7] or 0) for t in filtered if len(t)>7)
    total_balance = sum(float(t[16] or 0) for t in filtered if len(t)>16)

    return render_template("trips.html", trips=trips, filtered_trips=filtered,
                           settings=settings, headers=HEADERS,
                           trucks=trucks, drivers=drivers,
                           total_freight=total_freight, total_balance=total_balance)

@app.route("/new_trip", methods=["GET","POST"])
def new_trip():
    settings = load_settings()
    if request.method == "POST":
        d = request.form
        try:
            freight    = float(d.get("freight")    or 0)
            toll       = float(d.get("toll")       or 0)
            commission = float(d.get("commission") or 0)
            fuel_l     = float(d.get("fuel_liters")or 0)
            fuel_a     = float(d.get("fuel_amount")or 0)
            expenses   = float(d.get("expenses")   or 0)
            advance    = float(d.get("advance")    or 0)
            bill_amt   = float(d.get("bill_amount")or 0)
            total_trip = toll + commission + fuel_a + expenses + advance
            balance    = freight - total_trip - bill_amt
        except ValueError:
            return render_template("new_trip.html", settings=settings, error="Invalid numbers entered.")

        truck  = d.get("truck_no","").strip().upper()
        driver = d.get("driver_name","").strip().title()

        if truck and truck not in settings["trucks"]:
            settings["trucks"].append(truck)
        if driver and driver not in settings["drivers"]:
            settings["drivers"].append(driver)
        save_settings_file(settings)

        data = [
            d.get("date",""), truck, driver,
            d.get("loading_point","").strip().title(),
            d.get("delivery_point","").strip().title(),
            float(d.get("weight") or 0),
            freight, toll, commission, fuel_l, fuel_a, expenses, advance,
            bill_amt, round(total_trip,2), round(balance,2)
        ]
        save_trip_excel(data)
        return redirect(url_for("trips_view"))
    return render_template("new_trip.html", settings=settings, error=None,
                           today=datetime.now().strftime("%Y-%m-%d"))

@app.route("/edit_trip/<int:trip_no>", methods=["GET","POST"])
def edit_trip(trip_no):
    settings = load_settings()
    trips = load_trips()
    trip = next((t for t in trips if t[0] == trip_no), None)
    if not trip:
        return redirect(url_for("trips_view"))
    if request.method == "POST":
        d = request.form
        try:
            freight    = float(d.get("freight")    or 0)
            toll       = float(d.get("toll")       or 0)
            commission = float(d.get("commission") or 0)
            fuel_l     = float(d.get("fuel_liters")or 0)
            fuel_a     = float(d.get("fuel_amount")or 0)
            expenses   = float(d.get("expenses")   or 0)
            advance    = float(d.get("advance")    or 0)
            bill_amt   = float(d.get("bill_amount")or 0)
            total_trip = toll + commission + fuel_a + expenses + advance
            balance    = freight - total_trip - bill_amt
        except ValueError:
            return render_template("edit_trip.html", trip=trip, settings=settings, error="Invalid numbers.")
        data = [
            d.get("date",""),
            d.get("truck_no","").strip().upper(),
            d.get("driver_name","").strip().title(),
            d.get("loading_point","").strip().title(),
            d.get("delivery_point","").strip().title(),
            float(d.get("weight") or 0),
            freight, toll, commission, fuel_l, fuel_a, expenses, advance,
            bill_amt, round(total_trip,2), round(balance,2)
        ]
        update_trip_excel(trip_no, data)
        return redirect(url_for("trips_view"))
    return render_template("edit_trip.html", trip=trip, settings=settings, error=None)

@app.route("/delete_trip/<int:trip_no>", methods=["POST"])
def delete_trip(trip_no):
    delete_trip_excel(trip_no)
    return redirect(url_for("trips_view"))

@app.route("/dashboard")
def dashboard():
    settings = load_settings()
    trips = load_trips()
    total_trips   = len(trips)
    total_freight = sum(float(t[7] or 0) for t in trips if len(t)>7)
    total_balance = sum(float(t[16] or 0) for t in trips if len(t)>16)
    total_fuel    = sum(float(t[10] or 0) for t in trips if len(t)>10)
    trucks_count  = len(set(t[2] for t in trips if t[2]))

    truck_freight = {}
    for t in trips:
        k = str(t[2] or "Unknown")
        truck_freight[k] = truck_freight.get(k,0) + (float(t[7] or 0) if len(t)>7 else 0)

    monthly = {}
    for t in trips:
        try:
            m = datetime.strptime(str(t[1]), "%d-%m-%Y").strftime("%b %Y")
            monthly[m] = monthly.get(m,0) + 1
        except: pass

    driver_bal = {}
    for t in trips:
        k = str(t[3] or "Unknown")
        driver_bal[k] = driver_bal.get(k,0) + (float(t[16] or 0) if len(t)>16 else 0)

    return render_template("dashboard.html", settings=settings,
        total_trips=total_trips, total_freight=total_freight,
        total_balance=total_balance, total_fuel=total_fuel,
        trucks_count=trucks_count,
        truck_freight=truck_freight, monthly=monthly, driver_bal=driver_bal)

@app.route("/reports")
def reports():
    settings = load_settings()
    trips = load_trips()
    trucks  = sorted(set(str(t[2]) for t in trips if t[2]))
    drivers = sorted(set(str(t[3]) for t in trips if t[3]))
    return render_template("reports.html", settings=settings, trucks=trucks, drivers=drivers)

@app.route("/download_report", methods=["POST"])
def download_report():
    trips = load_trips()
    rtype  = request.form.get("report_type","full")
    truck  = request.form.get("truck","All")
    driver = request.form.get("driver","All")
    date_from = request.form.get("date_from","")
    date_to   = request.form.get("date_to","")

    filtered = []
    for t in trips:
        if truck != "All" and str(t[2]).upper() != truck.upper(): continue
        if driver != "All" and str(t[3]).lower() != driver.lower(): continue
        if date_from and date_to:
            try:
                td = datetime.strptime(str(t[1]), "%d-%m-%Y").date()
                fd = datetime.strptime(date_from, "%Y-%m-%d").date()
                tod= datetime.strptime(date_to,   "%Y-%m-%d").date()
                if not (fd <= td <= tod): continue
            except: pass
        filtered.append(t)

    if rtype in ("per_truck", "per_driver", "monthly"):
        wb = openpyxl.Workbook()
        first = True
        if rtype == "per_truck":
            groups = {}
            for t in filtered:
                groups.setdefault(str(t[2] or "Unknown"), []).append(t)
            label = "per_truck"
        elif rtype == "per_driver":
            groups = {}
            for t in filtered:
                groups.setdefault(str(t[3] or "Unknown"), []).append(t)
            label = "per_driver"
        else:
            groups = {}
            for t in filtered:
                try: m = datetime.strptime(str(t[1]), "%d-%m-%Y").strftime("%b %Y")
                except: m = "Unknown"
                groups.setdefault(m, []).append(t)
            label = "monthly"
        for name, tlist in sorted(groups.items()):
            tmp_wb = build_report_wb(tlist, f"- {name}")
            ws_src = tmp_wb.active
            if first:
                ws = wb.active; ws.title = name[:31]; first = False
            else:
                ws = wb.create_sheet(title=name[:31])
            for row in ws_src.iter_rows(values_only=True):
                ws.append(list(row))
        fname = f"report_{label}.xlsx"
    else:
        wb = build_report_wb(filtered, f"({rtype})")
        fname = f"report_{rtype}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/settings", methods=["GET","POST"])
def settings_view():
    settings = load_settings()
    msg = None
    if request.method == "POST":
        settings["company_name"]    = request.form.get("company_name","").strip()
        settings["company_address"] = request.form.get("company_address","").strip()
        settings["company_phone"]   = request.form.get("company_phone","").strip()
        trucks  = [t.strip() for t in request.form.get("trucks","").splitlines() if t.strip()]
        drivers = [d.strip() for d in request.form.get("drivers","").splitlines() if d.strip()]
        settings["trucks"]  = trucks
        settings["drivers"] = drivers
        save_settings_file(settings)
        msg = "Settings saved successfully."
    return render_template("settings.html", settings=settings, msg=msg)

@app.route("/backup", methods=["POST"])
def backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(BACKUP_DIR, f"truck_trips_backup_{ts}.xlsx")
    shutil.copy2(EXCEL_FILE, dest)
    return jsonify({"status": "ok", "file": os.path.basename(dest)})

@app.route("/api/calculate", methods=["POST"])
def calculate():
    d = request.json
    try:
        freight    = float(d.get("freight")    or 0)
        toll       = float(d.get("toll")       or 0)
        commission = float(d.get("commission") or 0)
        fuel_a     = float(d.get("fuel_amount")or 0)
        expenses   = float(d.get("expenses")   or 0)
        advance    = float(d.get("advance")    or 0)
        bill_amt   = float(d.get("bill_amount")or 0)
        total_trip = toll + commission + fuel_a + expenses + advance
        balance    = freight - total_trip - bill_amt
        return jsonify({"total_trip": round(total_trip,2), "balance": round(balance,2)})
    except: return jsonify({"total_trip": 0, "balance": 0})

if __name__ == "__main__":
    app.run(debug=True, port=5000)
