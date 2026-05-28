import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADERS = [
    "ID", "Date", "Truck No", "Driver Name", "Loading Point", "Delivery Point",
    "Weight (Tons)", "Freight", "Toll", "Commission", "Fuel Liters", "Fuel Amount",
    "Expenses", "Advance", "Bill Amount", "Total Trip Amount", "Balance Amount"
]
KEYS = [
    "id","date","truck_no","driver_name","loading_point","delivery_point",
    "weight","freight","toll","commission","fuel_liters","fuel_amount",
    "expenses","advance","bill_amount","total_trip_amount","balance_amount"
]

def write_sheet(ws, trips, subtitle, company):
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_col = get_column_letter(len(HEADERS))

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{company} - {subtitle}"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=10)

    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[4].height = 32

    alt = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    totals = {k: 0 for k in ["freight","toll","commission","fuel_liters","fuel_amount",
                               "expenses","advance","bill_amount","total_trip_amount","balance_amount"]}
    for i, trip in enumerate(trips, 1):
        rn = i + 4
        for col, key in enumerate(KEYS, 1):
            val = trip.get(key)
            cell = ws.cell(row=rn, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0: cell.fill = alt
        for k in totals:
            try: totals[k] += float(trip.get(k) or 0)
            except: pass

    tr = len(trips) + 5
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="center")
    tfill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    total_keys = ["freight","toll","commission","fuel_liters","fuel_amount",
                  "expenses","advance","bill_amount","total_trip_amount","balance_amount"]
    for col_offset, k in enumerate(total_keys):
        col = KEYS.index(k) + 1
        cell = ws.cell(row=tr, column=col, value=round(totals[k], 2))
        cell.font = Font(bold=True); cell.fill = tfill
        cell.border = border; cell.alignment = Alignment(horizontal="center")

    col_widths = [6,12,12,15,18,18,10,12,10,12,10,12,12,12,12,16,14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def build_report(trips, report_type, settings):
    company = settings.get("company_name", "GOODS CARRIER")
    wb = openpyxl.Workbook()

    if report_type in ("full", "filtered"):
        ws = wb.active
        ws.title = "Summary"
        write_sheet(ws, trips, "TRIP SUMMARY REPORT", company)

    elif report_type == "per_truck":
        groups = {}
        for t in trips: groups.setdefault(t["truck_no"] or "Unknown", []).append(t)
        first = True
        for name, tlist in sorted(groups.items()):
            ws = wb.active if first else wb.create_sheet(title=name[:31])
            if first: ws.title = name[:31]; first = False
            write_sheet(ws, tlist, f"Truck: {name}", company)

    elif report_type == "per_driver":
        groups = {}
        for t in trips: groups.setdefault(t["driver_name"] or "Unknown", []).append(t)
        first = True
        for name, tlist in sorted(groups.items()):
            ws = wb.active if first else wb.create_sheet(title=name[:31])
            if first: ws.title = name[:31]; first = False
            write_sheet(ws, tlist, f"Driver: {name}", company)

    elif report_type == "monthly":
        groups = {}
        for t in trips:
            try: m = datetime.strptime(t["date"], "%d-%m-%Y").strftime("%b %Y")
            except: m = "Unknown"
            groups.setdefault(m, []).append(t)
        first = True
        for name, tlist in sorted(groups.items()):
            ws = wb.active if first else wb.create_sheet(title=name[:31])
            if first: ws.title = name[:31]; first = False
            write_sheet(ws, tlist, f"Month: {name}", company)

    return wb
